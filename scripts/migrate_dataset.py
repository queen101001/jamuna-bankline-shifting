"""
One-time data migration: fix Left Bank coordinate inversion.

Reads dataset.xlsx from the project root, multiplies all "Left Bank (m)"
values by -1, and saves the corrected file to dataset/proper-dataset.xlsx.
The original is copied to dataset/dataset.xlsx as a backup.

Usage:
    python scripts/migrate_dataset.py
"""

from __future__ import annotations

import re
import shutil
from pathlib import Path

import openpyxl


PROJECT_ROOT = Path(__file__).resolve().parent.parent
SOURCE = PROJECT_ROOT / "dataset.xlsx"
DEST_DIR = PROJECT_ROOT / "dataset"
BACKUP = DEST_DIR / "dataset.xlsx"
OUTPUT = DEST_DIR / "proper-dataset.xlsx"

_LEFT_BANK_RE = re.compile(r"Left Bank", re.IGNORECASE)


def migrate() -> None:
    if not SOURCE.exists():
        raise FileNotFoundError(f"Source file not found: {SOURCE}")

    DEST_DIR.mkdir(parents=True, exist_ok=True)

    # Backup original
    shutil.copy2(SOURCE, BACKUP)
    print(f"Backed up original to {BACKUP}")

    # Load workbook
    wb = openpyxl.load_workbook(SOURCE)
    ws = wb["Data"]

    # Row 2 contains bank-side sub-headers ("Right Bank (m)", "Left Bank (m)")
    # Identify which columns are Left Bank
    left_bank_cols: list[int] = []
    for col_idx in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=2, column=col_idx).value
        if cell_val and _LEFT_BANK_RE.search(str(cell_val)):
            left_bank_cols.append(col_idx)

    print(f"Found {len(left_bank_cols)} Left Bank columns to invert")

    # Multiply Left Bank values by -1 (data rows start at row 3)
    flipped = 0
    for col_idx in left_bank_cols:
        for row_idx in range(3, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None and isinstance(cell.value, (int, float)):
                cell.value = cell.value * -1
                flipped += 1

    print(f"Flipped {flipped} Left Bank values (×-1)")

    # Save corrected file
    wb.save(OUTPUT)
    print(f"Saved corrected dataset to {OUTPUT}")

    # Verify
    wb2 = openpyxl.load_workbook(OUTPUT)
    ws2 = wb2["Data"]
    sample_orig = openpyxl.load_workbook(BACKUP)["Data"]

    # Show a few examples for verification
    print("\nVerification (first Left Bank column, reaches 1-5):")
    col = left_bank_cols[0]
    header = ws2.cell(row=1, column=col).value
    sub_header = ws2.cell(row=2, column=col).value
    print(f"  Column: {header} / {sub_header}")
    for row in range(3, 8):
        orig_val = sample_orig.cell(row=row, column=col).value
        new_val = ws2.cell(row=row, column=col).value
        reach = ws2.cell(row=row, column=1).value
        print(f"  Reach {reach}: {orig_val} → {new_val}")

    print("\nMigration complete!")


if __name__ == "__main__":
    migrate()
